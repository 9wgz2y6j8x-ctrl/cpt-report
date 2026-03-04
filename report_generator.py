"""
report_generator.py

Module de generation de rapports Excel et PDF pour les essais CPT.

Produit un fichier Excel par numero de dossier, avec une feuille par essai,
et un fichier PDF avec une page par essai (mise en page ReportLab).

Colonnes remplies :
  - Prof.  : profondeur reechantillonnee
  - Cote   : cote de depart - profondeur
  - qc     : resistance a la pointe corrigee [kg/cm2]
  - q'0    : contrainte naturelle effective [kg/cm2]
  - Qst    : frottement lateral total corrige [kg] (Excel uniquement)
  - phi'   : angle de frottement effectif [deg] (min 30 deg)
  - phi_u  : angle de frottement brut [deg]
  - Padm,1 : pression admissible sous semelle 1 [kg/cm2]
  - Padm,2 : pression admissible sous semelle 2 [kg/cm2]
  - C      : coefficient de compressibilite [-] (alpha * qc / q'0)
  - Nq     : facteur de portance Nq (propre a la methode choisie)
  - Ng     : facteur de portance Ngamma (Vpg pour De Beer / INISMa)

Les colonnes qc et Qst necessitent la selection d'une machine dans la
vue Calculer ; sans machine, elles restent vides (ainsi que phi', phi_u,
Padm1 et Padm2 qui dependent de qc).
"""

import os
import re
import math
import logging
from typing import List, Dict, Any, Optional, Callable

import io
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from tabular_reader import load_cpt_dataframe
from cpt_plot import CPTPlotConfig, _resolve_column_name, plot_cpt
from cpt_correction import (
    ParamsAppareilCPT,
    calculer_qc_corrige,
    calculer_qst_corrige,
    _compter_tiges,
)
from units import qc_to_internal, qst_to_internal, internal_to_plot
from friction_angle import calculer_angles_frottement
from bearing_capacity import calculer_pressions_admissibles, calculer_nq, calculer_ng

logger = logging.getLogger(__name__)


# ──────── Colonnes du rapport ────────

REPORT_COLUMNS = [
    "Prof.",
    "Cote",
    "qc",
    "q'0",
    "Qst",
    "\u03c6'",       # φ'
    "\u03c6u",       # φu
    "Padm, 1",
    "Padm, 2",
    "C",
    "Nq",
    "N\u03B3",       # Nγ
]

REPORT_UNITS_TEMPLATE = [
    "[m]",
    "[m]",
    "[kg/cm\u00b2]",   # [kg/cm²]
    "[kg/cm\u00b2]",   # [kg/cm²]
    "[kg]",
    "[\u00b0]",         # [°]
    "[\u00b0]",         # [°]
    "[kg/cm\u00b2] B={b1}",   # [kg/cm²] B=<largeur1>  — sera formatte
    "[kg/cm\u00b2] B={b2}",   # [kg/cm²] B=<largeur2>  — sera formatte
    "[/] \u03b1={alpha}",      # [/] α=<valeur>  — sera formatte
    "[/]",
    "[/]",
]


# ──────── Sanitization des noms de feuilles Excel ────────

# Caracteres interdits dans les noms de feuilles Excel
_EXCEL_FORBIDDEN = re.compile(r'[\[\]:*?/\\]')
_MAX_SHEET_NAME = 31


def _sanitize_sheet_name(name: str) -> str:
    """Nettoie un nom pour qu'il soit valide comme nom de feuille Excel.

    - Remplace les caracteres interdits par '_'
    - Tronque a 31 caracteres
    - Supprime les espaces en debut/fin
    """
    name = _EXCEL_FORBIDDEN.sub("_", name).strip()
    if not name:
        name = "Essai"
    return name[:_MAX_SHEET_NAME]


def _deduplicate_sheet_names(names: List[str]) -> List[str]:
    """Garantit l'unicite des noms de feuilles.

    Si deux noms identiques existent, ajoute un suffixe (_2, _3, ...).
    Respecte la limite de 31 caracteres.
    """
    seen: Dict[str, int] = {}
    result = []

    for name in names:
        lower = name.lower()
        if lower in seen:
            seen[lower] += 1
            suffix = f"_{seen[lower]}"
            # Tronquer pour laisser la place au suffixe
            max_base = _MAX_SHEET_NAME - len(suffix)
            deduped = name[:max_base] + suffix
            result.append(deduped)
        else:
            seen[lower] = 1
            result.append(name)

    return result


# ──────── Reechantillonnage ────────

def _resample_depths(
    depths: np.ndarray,
    step_m: float,
    prof_arrondie: float,
) -> np.ndarray:
    """Produit la liste des profondeurs reechantillonnees.

    Parametres
    ----------
    depths : np.ndarray
        Profondeurs triees issues des donnees (brutes ou filtrees).
    step_m : float
        Pas de reechantillonnage en metres (ex. 0.20).
    prof_arrondie : float
        Derniere profondeur a inclure (arrondie depuis la vue Calculer).

    Retours
    -------
    np.ndarray
        Profondeurs selectionnees apres reechantillonnage.

    Strategie
    ---------
    - Genere une grille reguliere : 0.00, step_m, 2*step_m, ..., prof_arrondie.
    - Pour chaque point de la grille, selectionne la profondeur reelle la plus
      proche si elle est dans une tolerance de step_m/2.
    - Si le pas reel des donnees est plus grand que le pas demande (upsampling
      interdit), on conserve les profondeurs originales et on log un warning.
    - La derniere profondeur (prof_arrondie) est toujours incluse, meme si
      absente des donnees.
    """
    if len(depths) < 2:
        return depths

    # Detecter le pas reel median des donnees
    diffs = np.diff(depths)
    data_step = float(np.median(diffs))

    # Si le pas reel est plus grand que le pas demande -> pas de sous-echantillonnage possible
    if data_step > step_m * 1.5:
        logger.warning(
            "Pas reel des donnees (%.4f m) superieur au pas demande (%.4f m). "
            "Les profondeurs originales sont conservees.",
            data_step, step_m,
        )
        # Ajouter prof_arrondie si elle n'est pas deja presente
        if prof_arrondie is not None and (len(depths) == 0 or abs(depths[-1] - prof_arrondie) > 1e-6):
            return np.append(depths, prof_arrondie)
        return depths

    # Generer la grille cible
    n_steps = int(round(prof_arrondie / step_m))
    target_depths = np.round(np.arange(0, n_steps + 1) * step_m, 6)

    # S'assurer que prof_arrondie est le dernier point
    if len(target_depths) == 0 or abs(target_depths[-1] - prof_arrondie) > 1e-6:
        target_depths = np.append(target_depths, prof_arrondie)

    # Selectionner les profondeurs : pour chaque cible, prendre la plus proche
    # dans les donnees reelles (tolerance = step_m / 2)
    tolerance = step_m / 2.0
    selected = []

    for target in target_depths:
        dists = np.abs(depths - target)
        idx = np.argmin(dists)
        if dists[idx] <= tolerance:
            selected.append(float(depths[idx]))
        else:
            # Pas de donnee proche -> inclure la profondeur cible quand meme
            # (utile surtout pour prof_arrondie qui peut depasser les donnees)
            selected.append(float(target))

    # Garantir que la derniere profondeur est exactement prof_arrondie
    if len(selected) > 0 and prof_arrondie is not None:
        selected[-1] = float(prof_arrondie)

    return np.array(selected)


# ──────── Contrainte naturelle effective (q'0) ────────

_RHO_EAU = 1000.0  # Masse volumique de l'eau [kg/m³]


def _contrainte_effective_verticale(
    profondeur: float,
    rho_sec: float,
    rho_sat: float,
    niveau_nappe: Optional[float],
) -> float:
    """Calcule la contrainte effective verticale sigma'_v a une profondeur donnee.

    Parametres
    ----------
    profondeur : float
        Profondeur par rapport au terrain naturel [m]. Doit etre >= 0.
    rho_sec : float
        Masse volumique du sol au-dessus de la nappe [kg/m³].
    rho_sat : float
        Masse volumique du sol sature en-dessous de la nappe [kg/m³].
    niveau_nappe : float | None
        Profondeur de la nappe [m]. None = pas de nappe.

    Retours
    -------
    float
        Contrainte effective verticale [kgf/cm²].

    Principe :
      - Au-dessus de la nappe (ou sans nappe) : sigma'_v = z * rho_sec / 10 000
      - En-dessous de la nappe :
            sigma'_v = [z_nappe * rho_sec + (z - z_nappe) * (rho_sat - rho_eau)] / 10 000
    """
    if profondeur <= 0:
        return 0.0

    nappe_presente = niveau_nappe is not None and niveau_nappe >= 0
    sous_la_nappe = nappe_presente and (profondeur >= niveau_nappe)

    if sous_la_nappe:
        z_nappe = niveau_nappe
        pression_au_dessus = z_nappe * rho_sec
        rho_dejauge = rho_sat - _RHO_EAU
        pression_en_dessous = (profondeur - z_nappe) * rho_dejauge
        return (pression_au_dessus + pression_en_dessous) / 10_000
    else:
        return (profondeur * rho_sec) / 10_000


def _resolve_niveau_nappe(observations: Optional[dict]) -> Optional[float]:
    """Determine le niveau de nappe a utiliser pour un essai.

    Priorite : fin_chantier > fin_essai > None (sans nappe).

    Parametres
    ----------
    observations : dict | None
        Donnees du store d'observations pour un essai.
        Structure attendue : {"hole_obs": {"Niveau d'eau": {"fin_essai": str, "fin_chantier": str}}, ...}

    Retours
    -------
    float | None
        Profondeur de la nappe [m], ou None si aucune donnee valide.
    """
    if not observations:
        return None

    hole_obs = observations.get("hole_obs", {})
    niveau_eau = hole_obs.get("Niveau d'eau", {})

    # Priorite 1 : fin de chantier
    fin_chantier = niveau_eau.get("fin_chantier", "").strip()
    if fin_chantier:
        try:
            val = float(fin_chantier.replace(",", "."))
            if val >= 0:
                return val
        except (ValueError, TypeError):
            pass

    # Priorite 2 : fin d'essai
    fin_essai = niveau_eau.get("fin_essai", "").strip()
    if fin_essai:
        try:
            val = float(fin_essai.replace(",", "."))
            if val >= 0:
                return val
        except (ValueError, TypeError):
            pass

    return None


# ──────── Recuperation des donnees filtrables ────────

def _get_dataframe_for_essai(
    file_path: str,
    file_data: dict,
    cleaning_entries: Optional[Dict[str, Any]] = None,
) -> Optional[pd.DataFrame]:
    """Retourne le DataFrame a utiliser pour un essai.

    Utilise les donnees filtrees si disponibles, sinon les donnees brutes.

    Parametres
    ----------
    file_path : str
        Chemin du fichier d'essai.
    file_data : dict
        Donnees du fichier (pour load_cpt_dataframe).
    cleaning_entries : dict, optional
        Mapping {file_path: CPTFileEntry} depuis la vue Filtrer.

    Retours
    -------
    pd.DataFrame ou None
    """
    # Verifier si des donnees filtrees existent
    if cleaning_entries and file_path in cleaning_entries:
        entry = cleaning_entries[file_path]
        if entry.is_filtered and entry.df_filtered is not None:
            logger.info("Utilisation des donnees filtrees pour %s", file_path)
            return entry.df_filtered

    # Sinon, charger les donnees brutes
    try:
        df = load_cpt_dataframe(file_data)
        if df.empty:
            logger.warning("DataFrame vide pour %s", file_path)
            return None
        return df
    except Exception as exc:
        logger.error("Erreur de chargement pour %s : %s", file_path, exc)
        return None


# ──────── Construction des parametres de correction ────────

def _build_correction_params(
    essai: Dict[str, Any],
    settings_manager,
    raw_data_manager=None,
) -> Optional[ParamsAppareilCPT]:
    """Construit les ParamsAppareilCPT depuis l'essai et la config machine.

    Retourne None si la machine n'est pas selectionnee ou introuvable.
    """
    machine_name = essai.get("machine", "").strip()
    if not machine_name:
        return None

    machines = settings_manager.get_machines()
    machine_cfg = next(
        (m for m in machines if m.get("nom") == machine_name), None
    )
    if machine_cfg is None:
        logger.warning(
            "Machine '%s' introuvable dans la configuration.", machine_name
        )
        return None

    section = essai.get("section", "Grande")
    if section == "Petite":
        poids_tige_kg = machine_cfg.get("poids_tige_petite_section", 0.0)
        poids_tube_kg = machine_cfg.get("poids_tube_petite_section", 0.0)
    else:
        poids_tige_kg = machine_cfg.get("poids_tige_grande_section", 0.0)
        poids_tube_kg = machine_cfg.get("poids_tube_grande_section", 0.0)

    tip_area_cm2 = settings_manager.get("unites", "tip_area_cm2") or 10.0
    section_pointe_m2 = tip_area_cm2 * 1e-4  # cm2 -> m2

    return ParamsAppareilCPT(
        section_pointe_m2=section_pointe_m2,
        poids_tige_kg=poids_tige_kg,
        poids_tube_kg=poids_tube_kg,
        delta_petit_mano_kg=essai.get("delta_petit", 0),
        delta_grand_mano_kg=essai.get("delta_grand", 0),
        nb_tubes_avant_sol=machine_cfg.get("nb_tubes_avant_sol", 0),
    )


def _compute_corrections(
    df: pd.DataFrame,
    col_depth: str,
    params: ParamsAppareilCPT,
    unit_qc: str,
    unit_qst: str,
    tip_area_cm2: float,
) -> Optional[Dict[str, np.ndarray]]:
    """Calcule qc corrige et Qst corrige pour un DataFrame d'essai.

    Retourne un dict avec les cles :
        'depths'   : profondeurs valides triees
        'qc_out'   : qc corrige en kg/cm2
        'qst_out'  : Qst corrige en kg

    Retourne None en cas d'erreur.
    """
    cfg = CPTPlotConfig()
    try:
        col_qc = _resolve_column_name(df, cfg.col_qc, "col_qc")
        col_qst = _resolve_column_name(df, cfg.col_qst, "col_qst")
    except Exception as exc:
        logger.warning("Colonnes qc/Qst introuvables : %s", exc)
        return None

    # Extraire les donnees valides (profondeur non NaN)
    mask = df[col_depth].notna()
    df_valid = df.loc[mask].copy()
    df_valid = df_valid.sort_values(by=col_depth).reset_index(drop=True)

    if df_valid.empty:
        return None

    valid_depths = df_valid[col_depth].values.astype(float)
    valid_qc = df_valid[col_qc].fillna(0).values.astype(float)
    valid_qst = df_valid[col_qst].fillna(0).values.astype(float)

    # Convertir en unites internes (DaN/m2 et DaN)
    qc_internal = qc_to_internal(valid_qc, unit_qc, tip_area_cm2)
    qst_internal = qst_to_internal(valid_qst, unit_qst)

    # Compter les tiges par profondeur
    depth_series = pd.Series(valid_depths)
    qc_series = pd.Series(qc_internal)
    qst_series = pd.Series(qst_internal)

    n_tiges = _compter_tiges(depth_series, params.nb_tubes_avant_sol)

    # Calculer qc corrige
    qc_corrige = calculer_qc_corrige(qc_series, n_tiges, params)

    # Reconstruire rtotale pour le calcul de Qst
    rtotale_danm2 = qc_series + qst_series / params.section_pointe_m2

    # Calculer Qst corrige
    qst_corrige = calculer_qst_corrige(
        rtotale_danm2, qc_corrige, n_tiges, params
    )

    # Convertir en unites de sortie (kg/cm2 et kgf)
    qc_out, qst_out, _, _ = internal_to_plot(
        qc_corrige.values, qst_corrige.values,
        pair="kg_kg", tip_area_cm2=tip_area_cm2,
    )

    return {
        "depths": valid_depths,
        "qc_out": qc_out,
        "qst_out": qst_out,
    }


# ──────── Styles Excel ────────

_FONT_HEADER = Font(name="Arial", size=12, bold=True)
_FONT_DATA = Font(name="Courier New", size=11)
_FILL_HEADER = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
_BORDER_BOTTOM = Border(bottom=Side(style="thin", color="000000"))
_COL_WIDTH = 12.5
_NUM_COLS = 12


def _format_worksheet(ws) -> None:
    """Applique le formatage standard a une feuille de rapport.

    - Largeur de colonnes : 12,5 pour les 12 colonnes
    - Police : Courier New 11 partout, Arial 12 gras pour la ligne 1
    - Trame de fond bleu-gris pale sur les lignes 1 et 2
    - Bordure inferieure entre la ligne 2 et la ligne 3
    """
    # Largeur des colonnes
    for col_idx in range(1, _NUM_COLS + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTH

    # Ligne 1 : Arial 12 gras + fond bleu-gris
    for col_idx in range(1, _NUM_COLS + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER

    # Ligne 2 : Courier New 11 + fond bleu-gris + bordure inferieure
    for col_idx in range(1, _NUM_COLS + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = _FONT_DATA
        cell.fill = _FILL_HEADER
        cell.border = _BORDER_BOTTOM

    # Lignes de donnees : Courier New 11 + format numerique
    # Colonnes 1 (Prof.) et 2 (Cote) : format 0.00
    # Toutes les autres colonnes : format 0.000
    for row in ws.iter_rows(min_row=3, max_col=_NUM_COLS):
        for cell in row:
            cell.font = _FONT_DATA
            if cell.value is not None:
                if cell.column in (1, 2):
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '0.000'


# ──────── Generation Excel ────────

def generate_excel_reports(
    essais: List[Dict[str, Any]],
    settings_manager,
    cleaning_entries: Optional[Dict[str, Any]] = None,
    raw_data_manager=None,
    cotes: Optional[Dict[str, float]] = None,
    observations: Optional[Dict[str, dict]] = None,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> Dict[str, str]:
    """Genere les fichiers Excel de rapport, un par numero de dossier.

    Parametres
    ----------
    essais : list[dict]
        Liste des essais depuis TraiterView.get_ordered_essais().
        Chaque dict contient : file_path, job, test, alpha, prof_arrondie, ...
    settings_manager : SettingsManager
        Gestionnaire de reglages (pour le pas de reechantillonnage et le dossier
        de sortie).
    cleaning_entries : dict, optional
        Mapping {file_path: CPTFileEntry} depuis la vue Filtrer pour acceder
        aux donnees filtrees.
    raw_data_manager : RawDataManager, optional
        Gestionnaire des donnees brutes (pour acceder aux file_data complets).
    cotes : dict, optional
        Mapping {file_path: cote_de_depart} depuis la vue Cotes.
        Si absent ou si un essai n'a pas de cote, cote_de_depart = 0.
    observations : dict, optional
        Mapping {file_path: store_dict} depuis la vue Observations.
        Chaque store_dict contient les niveaux d'eau (hole_obs) et annotations.
    progress_callback : callable, optional
        Fonction(current, total, message) pour le suivi de progression.

    Retours
    -------
    dict
        Mapping {numero_dossier: chemin_fichier_excel} des fichiers generes.

    Raises
    ------
    ValueError
        Si le dossier de sortie n'est pas configure.
    """
    # Recuperer les reglages
    dossier_resultats = settings_manager.get("dossiers_travail", "dossier_resultats")
    if not dossier_resultats or not dossier_resultats.strip():
        raise ValueError(
            "Le dossier d'enregistrement des resultats n'est pas configure. "
            "Veuillez le definir dans Reglages > Dossiers de travail."
        )

    step_cm = settings_manager.get("rapport", "reechantillonnage_cm") or 20
    step_m = step_cm / 100.0

    # Masses volumiques depuis les reglages
    rho_sec = settings_manager.get("parametres_calcul", "masse_volumique_sol_sec") or 1800
    rho_sat = settings_manager.get("parametres_calcul", "masse_volumique_sol_sature") or 2000

    # Parametres de calcul de portance
    methode_portance = (
        settings_manager.get("parametres_calcul", "methode_calcul_portance")
        or "De Beer (adapté)"
    )
    largeur_semelle_1 = (
        settings_manager.get("parametres_calcul", "largeur_semelle_fondation_1")
        or 0.6
    )
    largeur_semelle_2 = (
        settings_manager.get("parametres_calcul", "largeur_semelle_fondation_2")
        or 1.5
    )
    coeff_securite = (
        settings_manager.get("parametres_calcul", "coefficient_securite")
        or 2
    )

    # Grouper les essais par numero de dossier
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for essai in essais:
        job = essai.get("job", "").strip() or "Sans dossier"
        groups.setdefault(job, []).append(essai)

    total_essais = len(essais)
    current = 0
    generated_files: Dict[str, str] = {}

    for job_number, job_essais in groups.items():
        # Nom du fichier Excel
        safe_job = re.sub(r'[<>:"/\\|?*]', '_', job_number)
        filename = f"{safe_job}-Annexe resultats CPT.xlsx"
        filepath = os.path.join(dossier_resultats, filename)

        # Creer le classeur
        wb = Workbook()
        # Supprimer la feuille par defaut
        wb.remove(wb.active)

        # Preparer les noms de feuilles
        raw_names = []
        for essai in job_essais:
            test_name = essai.get("test", "").strip() or "Essai"
            raw_names.append(_sanitize_sheet_name(test_name))

        sheet_names = _deduplicate_sheet_names(raw_names)

        for i, essai in enumerate(job_essais):
            current += 1
            sheet_name = sheet_names[i]

            if progress_callback:
                progress_callback(
                    current, total_essais,
                    f"Dossier {job_number} - {sheet_name}"
                )

            ws = wb.create_sheet(title=sheet_name)

            # Ecrire les en-tetes (ligne 1 : titres)
            for col_idx, col_title in enumerate(REPORT_COLUMNS, start=1):
                ws.cell(row=1, column=col_idx, value=col_title)

            # Ecrire les unites (ligne 2)
            # Formater alpha (compressibilite) : "1.5" -> "1,5", "2.0" -> "2" (notation francaise)
            alpha_essai = essai.get("alpha", 1.5)
            alpha_str = (
                f"{alpha_essai:.1f}".replace(".", ",")
                if alpha_essai != int(alpha_essai)
                else str(int(alpha_essai))
            )
            # Formater les largeurs de semelles en cm pour l'affichage
            b1_cm = int(round(largeur_semelle_1 * 100))
            b2_cm = int(round(largeur_semelle_2 * 100))

            for col_idx, unit_template in enumerate(REPORT_UNITS_TEMPLATE, start=1):
                unit = unit_template.format(
                    alpha=alpha_str, b1=f"{b1_cm}cm", b2=f"{b2_cm}cm",
                )
                ws.cell(row=2, column=col_idx, value=unit)

            # Charger les donnees
            file_path = essai.get("file_path", "")
            file_data = None
            if raw_data_manager:
                file_data = raw_data_manager.get_file(file_path)
            if file_data is None:
                file_data = {"file_path": file_path}

            df = _get_dataframe_for_essai(file_path, file_data, cleaning_entries)
            if df is None:
                logger.warning(
                    "Impossible de charger les donnees pour %s, feuille laissee vide.",
                    file_path,
                )
                continue

            # Resoudre la colonne de profondeur
            try:
                cfg = CPTPlotConfig()
                col_depth = _resolve_column_name(df, cfg.col_depth, "col_depth")
            except Exception as exc:
                logger.error(
                    "Colonne de profondeur introuvable pour %s : %s",
                    file_path, exc,
                )
                continue

            # Extraire et trier les profondeurs
            depths = df[col_depth].dropna().values.astype(float)
            depths = np.sort(depths)

            # Profondeur arrondie (depuis la vue Calculer)
            prof_arrondie = essai.get("prof_arrondie")
            if prof_arrondie is None:
                prof_arrondie = float(depths[-1]) if len(depths) > 0 else 0.0

            # Reechantillonner
            resampled = _resample_depths(depths, step_m, prof_arrondie)

            # Omettre la premiere ligne a profondeur 0.00
            if len(resampled) > 0 and abs(resampled[0]) < 1e-6:
                resampled = resampled[1:]

            # Ecrire la colonne Profondeur (colonne 1, a partir de la ligne 3)
            for row_idx, depth_val in enumerate(resampled, start=3):
                ws.cell(row=row_idx, column=1, value=round(depth_val, 2))

            # Ecrire la colonne Cote (colonne 2) : cote = cote_de_depart - prof
            cote_depart = (cotes or {}).get(file_path, 0.0)
            for row_idx, depth_val in enumerate(resampled, start=3):
                ws.cell(row=row_idx, column=2, value=round(cote_depart - depth_val, 2))

            # ── Contrainte naturelle q'0 (colonne 4) ──
            obs_store = (observations or {}).get(file_path)
            niveau_nappe = _resolve_niveau_nappe(obs_store)

            q0_values = []
            for row_idx, depth_val in enumerate(resampled, start=3):
                q0 = _contrainte_effective_verticale(
                    depth_val, rho_sec, rho_sat, niveau_nappe,
                )
                ws.cell(row=row_idx, column=4, value=round(q0, 3))
                q0_values.append(q0)

            # ── Correction qc et Qst ──
            correction_params = _build_correction_params(
                essai, settings_manager, raw_data_manager,
            )

            if correction_params is not None:
                tip_area_cm2 = (
                    settings_manager.get("unites", "tip_area_cm2") or 10.0
                )
                unit_qc = "MPa"
                unit_qst = "kN"
                if raw_data_manager:
                    unit_qc = raw_data_manager.get_unit(file_path, "unit_qc")
                    unit_qst = raw_data_manager.get_unit(file_path, "unit_qst")

                corrections = _compute_corrections(
                    df, col_depth, correction_params,
                    unit_qc, unit_qst, tip_area_cm2,
                )

                if corrections is not None:
                    corr_depths = corrections["depths"]
                    qc_out = corrections["qc_out"]
                    qst_out = corrections["qst_out"]

                    # Pour chaque profondeur reechantillonnee, trouver la
                    # valeur corrigee la plus proche
                    for row_idx, depth_val in enumerate(resampled, start=3):
                        idx = int(np.argmin(np.abs(corr_depths - depth_val)))
                        qc_val = float(qc_out[idx])
                        ws.cell(
                            row=row_idx, column=3,
                            value=round(qc_val, 3),
                        )
                        ws.cell(
                            row=row_idx, column=5,
                            value=round(float(qst_out[idx]), 3),
                        )

                        # ── Angles de frottement phi' (col 6) et phi_u (col 7) ──
                        q0_val = q0_values[row_idx - 3]
                        phi_prime, phi_u = calculer_angles_frottement(qc_val, q0_val)
                        if phi_prime is not None:
                            ws.cell(row=row_idx, column=6, value=round(phi_prime, 3))
                        if phi_u is not None:
                            ws.cell(row=row_idx, column=7, value=round(phi_u, 3))

                        # ── Coefficient de compressibilite C (col 10) ──
                        # C = alpha * vbd, ou vbd = qc / q'0
                        if q0_val > 0:
                            vbd = qc_val / q0_val
                            coeff_c = alpha_essai * vbd
                            ws.cell(row=row_idx, column=10, value=round(coeff_c, 3))

                        # ── Pressions admissibles Padm1 (col 8) et Padm2 (col 9) ──
                        if phi_prime is not None and phi_u is not None and q0_val > 0:
                            padm1, padm2 = calculer_pressions_admissibles(
                                methode=methode_portance,
                                profondeur=depth_val,
                                q0_kgcm2=q0_val,
                                phip_deg=phi_prime,
                                phiu_deg=phi_u,
                                largeur_semelle_1_m=largeur_semelle_1,
                                largeur_semelle_2_m=largeur_semelle_2,
                                coeff_securite=coeff_securite,
                                rho_sec=rho_sec,
                                rho_sat=rho_sat,
                                niveau_nappe=niveau_nappe,
                                qc_kgcm2=qc_val,
                            )
                            ws.cell(row=row_idx, column=8, value=round(padm1, 3))
                            ws.cell(row=row_idx, column=9, value=round(padm2, 3))

                            # ── Nq (col 11) et Nγ (col 12) ──
                            nq_val = calculer_nq(
                                methode=methode_portance,
                                phiu_deg=phi_u,
                                phip_deg=phi_prime,
                                q0_kgcm2=q0_val,
                            )
                            ng_val = calculer_ng(
                                methode=methode_portance,
                                phiu_deg=phi_u,
                            )
                            ws.cell(row=row_idx, column=11, value=round(nq_val, 3))
                            ws.cell(row=row_idx, column=12, value=round(ng_val, 3))
            else:
                machine_name = essai.get("machine", "").strip()
                if not machine_name:
                    logger.info(
                        "Aucune machine selectionnee pour %s, colonnes qc/Qst "
                        "laissees vides.",
                        file_path,
                    )

            # ── Formatage de la feuille ──
            _format_worksheet(ws)

        # Sauvegarder le classeur
        try:
            os.makedirs(dossier_resultats, exist_ok=True)
            wb.save(filepath)
            generated_files[job_number] = filepath
            logger.info("Fichier Excel genere : %s", filepath)
        except OSError as exc:
            logger.error(
                "Erreur d'ecriture du fichier %s : %s", filepath, exc
            )
            raise

    return generated_files


# ══════════════════════════════════════════════════════════════════════
# Calcul des donnees partagees (utilise par Excel et PDF)
# ══════════════════════════════════════════════════════════════════════

def _compute_essai_data(
    essai: Dict[str, Any],
    settings_manager,
    cleaning_entries: Optional[Dict[str, Any]],
    raw_data_manager,
    cotes: Optional[Dict[str, float]],
    observations: Optional[Dict[str, dict]],
) -> Optional[Dict[str, Any]]:
    """Calcule toutes les donnees d'un essai pour le rapport.

    Retourne un dict contenant :
        'resampled'    : np.ndarray des profondeurs reechantillonnees (sans 0.00)
        'cote_depart'  : float
        'q0_values'    : list[float]
        'qc_values'    : list[float | None]
        'qst_values'   : list[float | None]
        'phi_prime_values' : list[float | None]
        'phi_u_values'     : list[float | None]
        'padm1_values'     : list[float | None]
        'padm2_values'     : list[float | None]
        'coeff_c_values'   : list[float | None]
        'nq_values'        : list[float | None]
        'ng_values'        : list[float | None]

    Retourne None si les donnees ne peuvent pas etre chargees.
    """
    step_cm = settings_manager.get("rapport", "reechantillonnage_cm") or 20
    step_m = step_cm / 100.0

    rho_sec = settings_manager.get("parametres_calcul", "masse_volumique_sol_sec") or 1800
    rho_sat = settings_manager.get("parametres_calcul", "masse_volumique_sol_sature") or 2000

    methode_portance = (
        settings_manager.get("parametres_calcul", "methode_calcul_portance")
        or "De Beer (adapté)"
    )
    largeur_semelle_1 = (
        settings_manager.get("parametres_calcul", "largeur_semelle_fondation_1")
        or 0.6
    )
    largeur_semelle_2 = (
        settings_manager.get("parametres_calcul", "largeur_semelle_fondation_2")
        or 1.5
    )
    coeff_securite = (
        settings_manager.get("parametres_calcul", "coefficient_securite")
        or 2
    )
    alpha_essai = essai.get("alpha", 1.5)

    file_path = essai.get("file_path", "")
    file_data = None
    if raw_data_manager:
        file_data = raw_data_manager.get_file(file_path)
    if file_data is None:
        file_data = {"file_path": file_path}

    df = _get_dataframe_for_essai(file_path, file_data, cleaning_entries)
    if df is None:
        return None

    try:
        cfg = CPTPlotConfig()
        col_depth = _resolve_column_name(df, cfg.col_depth, "col_depth")
    except Exception:
        return None

    depths = df[col_depth].dropna().values.astype(float)
    depths = np.sort(depths)

    prof_arrondie = essai.get("prof_arrondie")
    if prof_arrondie is None:
        prof_arrondie = float(depths[-1]) if len(depths) > 0 else 0.0

    resampled = _resample_depths(depths, step_m, prof_arrondie)

    # Omettre la premiere ligne a profondeur 0.00
    if len(resampled) > 0 and abs(resampled[0]) < 1e-6:
        resampled = resampled[1:]

    cote_depart = (cotes or {}).get(file_path, 0.0)

    obs_store = (observations or {}).get(file_path)
    niveau_nappe = _resolve_niveau_nappe(obs_store)

    n = len(resampled)
    q0_values = []
    qc_values = [None] * n
    qst_values = [None] * n
    phi_prime_values = [None] * n
    phi_u_values = [None] * n
    padm1_values = [None] * n
    padm2_values = [None] * n
    coeff_c_values = [None] * n
    nq_values = [None] * n
    ng_values = [None] * n

    for i, depth_val in enumerate(resampled):
        q0 = _contrainte_effective_verticale(depth_val, rho_sec, rho_sat, niveau_nappe)
        q0_values.append(q0)

    correction_params = _build_correction_params(essai, settings_manager, raw_data_manager)

    if correction_params is not None:
        tip_area_cm2 = settings_manager.get("unites", "tip_area_cm2") or 10.0
        unit_qc = "MPa"
        unit_qst = "kN"
        if raw_data_manager:
            unit_qc = raw_data_manager.get_unit(file_path, "unit_qc")
            unit_qst = raw_data_manager.get_unit(file_path, "unit_qst")

        corrections = _compute_corrections(
            df, col_depth, correction_params, unit_qc, unit_qst, tip_area_cm2,
        )

        if corrections is not None:
            corr_depths = corrections["depths"]
            qc_out = corrections["qc_out"]
            qst_out = corrections["qst_out"]

            for i, depth_val in enumerate(resampled):
                idx = int(np.argmin(np.abs(corr_depths - depth_val)))
                qc_val = float(qc_out[idx])
                qc_values[i] = qc_val
                qst_values[i] = float(qst_out[idx])

                q0_val = q0_values[i]
                phi_prime, phi_u = calculer_angles_frottement(qc_val, q0_val)
                phi_prime_values[i] = phi_prime
                phi_u_values[i] = phi_u

                if q0_val > 0:
                    vbd = qc_val / q0_val
                    coeff_c_values[i] = alpha_essai * vbd

                if phi_prime is not None and phi_u is not None and q0_val > 0:
                    padm1, padm2 = calculer_pressions_admissibles(
                        methode=methode_portance,
                        profondeur=depth_val,
                        q0_kgcm2=q0_val,
                        phip_deg=phi_prime,
                        phiu_deg=phi_u,
                        largeur_semelle_1_m=largeur_semelle_1,
                        largeur_semelle_2_m=largeur_semelle_2,
                        coeff_securite=coeff_securite,
                        rho_sec=rho_sec,
                        rho_sat=rho_sat,
                        niveau_nappe=niveau_nappe,
                        qc_kgcm2=qc_val,
                    )
                    padm1_values[i] = padm1
                    padm2_values[i] = padm2

                    nq_values[i] = calculer_nq(
                        methode=methode_portance,
                        phiu_deg=phi_u,
                        phip_deg=phi_prime,
                        q0_kgcm2=q0_val,
                    )
                    ng_values[i] = calculer_ng(
                        methode=methode_portance,
                        phiu_deg=phi_u,
                    )

    return {
        "resampled": resampled,
        "cote_depart": cote_depart,
        "q0_values": q0_values,
        "qc_values": qc_values,
        "qst_values": qst_values,
        "phi_prime_values": phi_prime_values,
        "phi_u_values": phi_u_values,
        "padm1_values": padm1_values,
        "padm2_values": padm2_values,
        "coeff_c_values": coeff_c_values,
        "nq_values": nq_values,
        "ng_values": ng_values,
    }


# ══════════════════════════════════════════════════════════════════════
# Generation PDF (ReportLab)
# ══════════════════════════════════════════════════════════════════════

def _init_pdf_fonts():
    """Enregistre les polices pour le rapport PDF.

    Calibri (Windows) / Carlito (Linux fallback), puis Arial.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    calibri_paths = {
        "Calibri":      ["C:/Windows/Fonts/calibri.ttf",
                         "/usr/share/fonts/truetype/crosextra/Carlito-Regular.ttf"],
        "Calibri-Bold": ["C:/Windows/Fonts/calibrib.ttf",
                         "/usr/share/fonts/truetype/crosextra/Carlito-Bold.ttf"],
    }
    for name, paths in calibri_paths.items():
        for p in paths:
            if os.path.exists(p):
                try:
                    pdfmetrics.registerFont(TTFont(name, p))
                    break
                except Exception:
                    continue

    arial_paths = {
        "Arial":      ["C:/Windows/Fonts/arial.ttf",
                       "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf"],
        "Arial-Bold": ["C:/Windows/Fonts/arialbd.ttf",
                       "/usr/share/fonts/truetype/msttcorefonts/Arial_Bold.ttf"],
    }
    for name, paths in arial_paths.items():
        for p in paths:
            if os.path.exists(p):
                try:
                    pdfmetrics.registerFont(TTFont(name, p))
                    break
                except Exception:
                    continue

    # Determiner les polices disponibles
    font_normal = "Calibri"
    font_bold = "Calibri-Bold"
    try:
        pdfmetrics.getFont(font_normal)
    except KeyError:
        font_normal = "Helvetica"
        font_bold = "Helvetica-Bold"

    try:
        pdfmetrics.getFont("Arial")
        font_arial = "Arial"
        font_arial_bold = "Arial-Bold"
    except KeyError:
        font_arial = font_normal
        font_arial_bold = font_bold

    return font_normal, font_bold, font_arial, font_arial_bold


# ──────── Constantes PDF ────────

_PDF_MAX_DATA_ROWS = 50

# Caractere phi
_PHI = "\u03C6"


def _pdf_draw_text_with_sub_super(c, x, y, fragments, font_name, font_size,
                                  alignment="center", col_width=0,
                                  font_bold=None):
    """Dessine du texte avec indices/exposants/gras sur le canvas PDF."""
    from reportlab.pdfbase import pdfmetrics as pm

    if font_bold is None:
        font_bold = font_name

    total_width = 0
    for text, style in fragments:
        sz = font_size * 0.65 if style in ("sub", "super") else font_size
        fn = font_bold if style == "bold" else font_name
        total_width += pm.stringWidth(text, fn, sz)

    if alignment == "center":
        draw_x = x + col_width / 2 - total_width / 2
    elif alignment == "right":
        draw_x = x + col_width - total_width - 2
    else:
        draw_x = x + 2

    for text, style in fragments:
        if style in ("sub", "super"):
            sz = font_size * 0.65
            fn = font_name
            dy = -font_size * 0.12 if style == "sub" else font_size * 0.35
            c.setFont(fn, sz)
            c.drawString(draw_x, y + dy, text)
            draw_x += pm.stringWidth(text, fn, sz)
        else:
            sz = font_size
            fn = font_bold if style == "bold" else font_name
            c.setFont(fn, sz)
            c.drawString(draw_x, y, text)
            draw_x += pm.stringWidth(text, fn, sz)


def _pdf_draw_footer_line_with_m3(c, x, y, prefix, font_name, font_size):
    """Dessine une ligne contenant [kg/m3] avec exposant."""
    from reportlab.pdfbase import pdfmetrics as pm

    c.setFont(font_name, font_size)
    c.drawString(x, y, prefix)
    w = pm.stringWidth(prefix, font_name, font_size)
    c.drawString(x + w, y, "[kg/m")
    w += pm.stringWidth("[kg/m", font_name, font_size)
    small_sz = font_size * 0.65
    c.setFont(font_name, small_sz)
    c.drawString(x + w, y + font_size * 0.35, "3")
    w += pm.stringWidth("3", font_name, small_sz)
    c.setFont(font_name, font_size)
    c.drawString(x + w, y, "]")


def _draw_diagram_footer(c, left_margin, bottom_margin, table_width,
                         font_normal, font_bold, observations, file_path,
                         essai, settings_manager, plot_pair):
    """Dessine le pied de page des pages diagrammes.

    Layout :
      Bloc gauche (~60 %) : Légende + Matériel utilisé (texte libre)
      Bloc droit  (~40 %) : Tableau Observations (3 cols × 3 lignes)

    Retourne la hauteur totale du bloc pied de page (en pts).
    """
    from reportlab.lib.colors import black
    from reportlab.pdfbase import pdfmetrics as pm
    from reportlab.lib.units import mm 

    # ── Constantes de dimensionnement ──
    FS = 7.5            # police contenu
    FS_LABEL = 8        # police titres de section (Légende, Matériel utilisé)
    LINE_H = 7.6         # hauteur d'une ligne de texte
    PAD = 2            # padding intérieur
    SECTION_GAP = 3     # espace vertical entre les deux sections du bloc gauche

    # Proportions gauche / droite
    LEFT_RATIO = 0.58
    left_w = table_width * LEFT_RATIO
    GAP_BETWEEN = 8     # espace blanc entre bloc gauche et tableau droit

    # ── Tableau droit : dimensions ──
    ROW_H = 15.5          # hauteur d'une ligne de données
    HDR_H = 25.5          # hauteur de l'en-tête (texte sur 2 lignes)
    tbl_h = HDR_H + ROW_H * 2

    # Hauteur totale du footer = max(bloc gauche, tableau droit)
    # Bloc gauche : Légende (label + 3 lignes) + gap + Matériel (label + 3 lignes)
    left_h = (LINE_H + 3 * LINE_H) + SECTION_GAP + (LINE_H + 3 * LINE_H)
    DIAG_FOOTER_HEIGHT = max(left_h, tbl_h) + 2  # petite marge

    # Coordonnées de base
    footer_top_y = bottom_margin + DIAG_FOOTER_HEIGHT
    tbl_x = left_margin + left_w + GAP_BETWEEN
    tbl_w = table_width - left_w - GAP_BETWEEN
    # Décalage vertical du tableau : positif = vers le bas, négatif = vers le haut
    OBS_TABLE_V_OFFSET = 2.8 * mm   # X mm vers le bas (changer le signe pour remonter)

    # ── BLOC GAUCHE ──
    cur_y = footer_top_y  # curseur vertical (descend)

    # --- Section Légende ---
    cur_y -= LINE_H
    c.setFont(font_bold, FS_LABEL)
    c.drawString(left_margin, cur_y, "Légende :")

    legend_items = [
        ("c",   "Chocs pendant l'enfoncement"),
        ("E",   "Extraction partielle des tubes de sondage"),
        ("TRF", "Sondage poursuivi avec Tube Réducteur de Frottement"),
    ]
    desc_x = left_margin + 30  # colonne des descriptions alignées
    c.setFont(font_normal, FS)
    for abbr, desc in legend_items:
        cur_y -= LINE_H
        c.drawString(left_margin + 4, cur_y, abbr)
        c.drawString(left_margin + 14, cur_y, "\u2192")
        c.drawString(desc_x, cur_y, desc)

    # --- Section Matériel utilisé ---
    cur_y -= SECTION_GAP + LINE_H
    c.setFont(font_bold, FS_LABEL)
    c.drawString(left_margin, cur_y, "Matériel utilisé :")

    # Récupérer les données machine
    machine_name = essai.get("machine", "").strip()
    capacite_str = "-"
    if machine_name:
        machines = settings_manager.get_machines()
        machine_cfg = next(
            (m for m in machines if m.get("nom") == machine_name), None)
        if machine_cfg:
            cap = machine_cfg.get("capacite_tonnes", 0)
            capacite_str = (f"{int(cap)}" if cap == int(cap)
                            else f"{cap}")

    section = essai.get("section", "Grande")
    section_pointe_cm2 = "10" if section == "Grande" else "6,6"
    section_type = "M1"
    section_tubes_cm2 = "10" if section == "Grande" else "6,6"

    # Position X pour les valeurs en gras (alignées à droite de la colonne)
    val_x = left_margin + left_w - 4

    c.setFont(font_normal, FS)
    # Ligne 1 : Capacité
    cur_y -= LINE_H
    c.drawString(left_margin + 4, cur_y, "Capacité de l'appareil hydraulique [T]")
    c.setFont(font_bold, FS)
    c.drawRightString(val_x, cur_y, capacite_str)

    # Ligne 2 : Section de la pointe [cm²] et type
    cur_y -= LINE_H
    c.setFont(font_normal, FS)
    lbl_part1 = "Section de la pointe [cm"
    c.drawString(left_margin + 4, cur_y, lbl_part1)
    x_after = left_margin + 4 + pm.stringWidth(lbl_part1, font_normal, FS)
    # Exposant ²
    c.setFont(font_normal, FS * 0.7)
    c.drawString(x_after, cur_y + 2.5, "2")
    x_after += pm.stringWidth("2", font_normal, FS * 0.7)
    c.setFont(font_normal, FS)
    c.drawString(x_after, cur_y, "] et type")
    # Valeur en gras : "10 M1"
    c.setFont(font_bold, FS)
    c.drawRightString(val_x, cur_y, f"{section_pointe_cm2} {section_type}")

    # Ligne 3 : Section des tubes allongés [cm²]
    cur_y -= LINE_H
    c.setFont(font_normal, FS)
    lbl_part1 = "Section des tubes allongés [cm"
    c.drawString(left_margin + 4, cur_y, lbl_part1)
    x_after = left_margin + 4 + pm.stringWidth(lbl_part1, font_normal, FS)
    c.setFont(font_normal, FS * 0.7)
    c.drawString(x_after, cur_y + 2.5, "2")
    x_after += pm.stringWidth("2", font_normal, FS * 0.7)
    c.setFont(font_normal, FS)
    c.drawString(x_after, cur_y, "]")
    c.setFont(font_bold, FS)
    c.drawRightString(val_x, cur_y, section_tubes_cm2)

    # ── BLOC DROIT : Tableau Observations ──
    c.setStrokeColor(black)
    c.setLineWidth(0.5)

    # Récupérer les données d'observation
    obs_store = None
    if observations and file_path in observations:
        obs_store = observations[file_path]

    def _fmt_obs(row_key, col_key):
        if obs_store is None:
            return "-"
        val_str = (obs_store.get("hole_obs", {})
                   .get(row_key, {})
                   .get(col_key, "").strip())
        if not val_str or val_str == "-":
            return "-"
        try:
            return f"{float(val_str.replace(',', '.')):.2f}".replace(".", ",")
        except (ValueError, TypeError):
            return val_str

    # Colonnes du tableau : col1 (observations), col2 (fin essai), col3 (fin chantier)
    col1_w = tbl_w * 0.34
    col2_w = tbl_w * 0.33
    col3_w = tbl_w * 0.33
    col1_x = tbl_x
    col2_x = tbl_x + col1_w
    col3_x = tbl_x + col1_w + col2_w

    # Le tableau est aligné en haut par rapport au bloc gauche
    tbl_top_y = footer_top_y - OBS_TABLE_V_OFFSET

    # Dessiner le cadre extérieur du tableau
    c.rect(tbl_x, tbl_top_y - tbl_h, tbl_w, tbl_h, stroke=1, fill=0)

    # Lignes verticales internes
    c.line(col2_x, tbl_top_y, col2_x, tbl_top_y - tbl_h)
    c.line(col3_x, tbl_top_y, col3_x, tbl_top_y - tbl_h)

    # Ligne horizontale sous l'en-tête
    hdr_bottom_y = tbl_top_y - HDR_H
    c.line(tbl_x, hdr_bottom_y, tbl_x + tbl_w, hdr_bottom_y)

    # Ligne horizontale entre les deux lignes de données
    r1_bottom_y = hdr_bottom_y - ROW_H
    c.line(tbl_x, r1_bottom_y, tbl_x + tbl_w, r1_bottom_y)

    # ── En-tête du tableau (texte sur 2 lignes) ──
    hdr_line1_y = tbl_top_y - 9     # première ligne d'en-tête
    hdr_line2_y = tbl_top_y - 18    # deuxième ligne d'en-tête

    c.setFont(font_normal, FS)

    # Col 1 : "Observations" / "dans le trou"
    c.drawCentredString(col1_x + col1_w / 2, hdr_line1_y, "Observations")
    c.drawCentredString(col1_x + col1_w / 2, hdr_line2_y, "dans le trou")

    # Col 2 : "Profondeur en" / "fin d'essai [m]"
    c.drawCentredString(col2_x + col2_w / 2, hdr_line1_y, "Profondeur en")
    c.drawCentredString(col2_x + col2_w / 2, hdr_line2_y, "fin d'essai [m]")

    # Col 3 : "Profondeur en" / "fin de chantier [m]"
    c.drawCentredString(col3_x + col3_w / 2, hdr_line1_y, "Profondeur en")
    c.drawCentredString(col3_x + col3_w / 2, hdr_line2_y, "fin de chantier [m]")

    # ── Ligne de données 1 : Niveau d'eau ──
    r1_text_y = hdr_bottom_y - ROW_H + PAD + (ROW_H - FS) / 2   # ← centrage vertical
    c.setFont(font_normal, FS)
    c.drawString(col1_x + PAD, r1_text_y, "Niveau d'eau")
    c.drawCentredString(col2_x + col2_w / 2, r1_text_y,
                        _fmt_obs("Niveau d'eau", "fin_essai"))
    c.drawCentredString(col3_x + col3_w / 2, r1_text_y,
                        _fmt_obs("Niveau d'eau", "fin_chantier"))

    # ── Ligne de données 2 : Eboulement ──
    r2_text_y = r1_bottom_y - ROW_H + PAD + (ROW_H - FS) / 2    # ← centrage vertica
    c.drawString(col1_x + PAD, r2_text_y, "Eboulement")
    c.drawCentredString(col2_x + col2_w / 2, r2_text_y,
                        _fmt_obs("Eboulement", "fin_essai"))
    c.drawCentredString(col3_x + col3_w / 2, r2_text_y,
                        _fmt_obs("Eboulement", "fin_chantier"))

    return DIAG_FOOTER_HEIGHT


def _format_date_for_pdf(date_str: str) -> str:
    """Convertit une date en format MM/AAAA pour le rapport PDF.

    Accepte divers formats d'entree et retourne MM/AAAA.
    """
    if not date_str or not date_str.strip():
        return ""
    date_str = date_str.strip()

    # Essayer differents formats courants
    import re as _re
    # Format DD/MM/YYYY ou DD-MM-YYYY
    m = _re.match(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})', date_str)
    if m:
        return f"{m.group(2)}/{m.group(3)}"
    # Format YYYY-MM-DD
    m = _re.match(r'(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', date_str)
    if m:
        return f"{m.group(2)}/{m.group(1)}"
    # Format MM/YYYY deja correct
    m = _re.match(r'(\d{1,2})[/\-.](\d{4})$', date_str)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    # Retourner tel quel si non reconnu
    return date_str


def _render_figure_to_image(fig, dpi=300):
    """Rend une figure matplotlib en image PNG en memoire.

    Parametres
    ----------
    fig : matplotlib.figure.Figure
        Figure a rendre.
    dpi : int
        Resolution en points par pouce.

    Retours
    -------
    io.BytesIO
        Buffer contenant l'image PNG.
    """
    import matplotlib.pyplot as plt

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return buf


def generate_pdf_report(
    essais: List[Dict[str, Any]],
    settings_manager,
    cleaning_entries: Optional[Dict[str, Any]] = None,
    raw_data_manager=None,
    cotes: Optional[Dict[str, float]] = None,
    observations: Optional[Dict[str, dict]] = None,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> Dict[str, str]:
    """Genere les fichiers PDF de rapport, un par numero de dossier.

    Chaque essai occupe une page du PDF, avec la mise en page ReportLab.

    Parametres
    ----------
    essais : list[dict]
        Liste des essais depuis TraiterView.get_ordered_essais().
    settings_manager : SettingsManager
        Gestionnaire de reglages.
    cleaning_entries : dict, optional
        Mapping {file_path: CPTFileEntry}.
    raw_data_manager : RawDataManager, optional
        Gestionnaire des donnees brutes.
    cotes : dict, optional
        Mapping {file_path: cote_de_depart}.
    observations : dict, optional
        Mapping {file_path: store_dict}.
    progress_callback : callable, optional
        Fonction(current, total, message).

    Retours
    -------
    dict
        Mapping {numero_dossier: chemin_fichier_pdf}.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import black
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics as pm

    font_normal, font_bold, font_arial, font_arial_bold = _init_pdf_fonts()

    dossier_resultats = settings_manager.get("dossiers_travail", "dossier_resultats")
    if not dossier_resultats or not dossier_resultats.strip():
        raise ValueError(
            "Le dossier d'enregistrement des resultats n'est pas configure."
        )

    rho_sec = settings_manager.get("parametres_calcul", "masse_volumique_sol_sec") or 1800
    rho_sat = settings_manager.get("parametres_calcul", "masse_volumique_sol_sature") or 2000
    coeff_securite = (
        settings_manager.get("parametres_calcul", "coefficient_securite") or 2
    )

    # ── Dimensions de page ──
    PAGE_W, PAGE_H = A4
    LEFT_MARGIN = 10 * mm
    RIGHT_MARGIN = 10 * mm
    TOP_MARGIN = 14 * mm
    BOTTOM_MARGIN = 8 * mm
    TABLE_WIDTH = PAGE_W - LEFT_MARGIN - RIGHT_MARGIN

    LOGO_LEFT = LEFT_MARGIN
    LOGO_WIDTH_PX = 145
    LOGO_HEIGHT_PX = 70
    LOGO_V_OFFSET = 15
    LABELS_LEFT = LEFT_MARGIN + LOGO_WIDTH_PX + 10

    META_BLOCK_WIDTH = 120
    META_X_LABEL = LEFT_MARGIN + TABLE_WIDTH - META_BLOCK_WIDTH
    META_X_VALUE = LEFT_MARGIN + TABLE_WIDTH - 8

    COL_RATIOS = [1.0, 1.0, 1.2, 1.2, 0.8, 0.8, 0.8, 0.8, 1.1, 1.1, 1.1]
    _total_ratio = sum(COL_RATIOS)
    COL_WIDTHS = [r / _total_ratio * TABLE_WIDTH for r in COL_RATIOS]

    ROW_HEIGHT = 12.25
    HEADER_ROW_HEIGHT = 17
    UNIT_ROW_HEIGHT = 16

    FONT_SIZE_COL_HEADER = 11
    FONT_SIZE_COL_UNIT = 11
    FONT_SIZE_DATA = 10
    FONT_SIZE_TITLE = 16
    FONT_SIZE_SUBTITLE = 11
    FONT_SIZE_TEST_TYPE = 10
    FONT_SIZE_SMALL = 7
    FONT_SIZE_META = 8
    FONT_SIZE_FOOTER = 10

    FOOTER_BOX_HEIGHT = 40

    def col_x(col_index):
        return LEFT_MARGIN + sum(COL_WIDTHS[:col_index])

    # Rechercher le logo
    logo_path = None
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for candidate in ["icons/Inisma.jpg", "icons/inisma.jpg", "icons/INISMA.jpg"]:
        full = os.path.join(script_dir, candidate)
        if os.path.exists(full):
            logo_path = full
            break

    # Texte societe (statique)
    company_line1 = "Géotechnique et Environnement Sol - 32(0)65/40 34 34 - Fax: 32(0)65/34 80 05"
    company_line2 = "Avenue Gouverneur Cornez 4, B-7000 Mons (Belgique) - www.bcrc.be"

    # ── Grouper par numero de dossier ──
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for essai in essais:
        job = essai.get("job", "").strip() or "Sans dossier"
        groups.setdefault(job, []).append(essai)

    total_essais = len(essais) * 2  # Pages diagrammes + pages tableaux
    current = 0
    generated_files: Dict[str, str] = {}

    # Paire d'unites graphique pour les diagrammes
    _plot_pair = settings_manager.get("unites", "paire_graphique") or "MPa_kN"

    for job_number, job_essais in groups.items():
        safe_job = re.sub(r'[<>:"/\\|?*]', '_', job_number)
        filename = f"{safe_job}-Rapport CPT.pdf"
        filepath = os.path.join(dossier_resultats, filename)

        os.makedirs(dossier_resultats, exist_ok=True)
        c = canvas.Canvas(filepath, pagesize=A4)
        c.setTitle("CPT Report - Sondage au Pénétromètre Statique")

        # ══════════════════════════════════════════════════════════════
        # Pages diagrammes (une page par essai, au debut du PDF)
        # ══════════════════════════════════════════════════════════════
        for essai in job_essais:
            current += 1
            test_name = essai.get("test", "").strip() or "Essai"

            if progress_callback:
                progress_callback(
                    current, total_essais,
                    f"Diagramme {job_number} - {test_name}"
                )

            file_path_diag = essai.get("file_path", "")
            file_data_diag = None
            if raw_data_manager:
                file_data_diag = raw_data_manager.get_file(file_path_diag)
            if file_data_diag is None:
                file_data_diag = {"file_path": file_path_diag}

            # Charger les donnees (filtrees si disponibles, brutes sinon)
            df_diag = _get_dataframe_for_essai(
                file_path_diag, file_data_diag, cleaning_entries,
            )

            if df_diag is None:
                c.showPage()
                continue

            # Resoudre les colonnes
            cfg_diag = CPTPlotConfig()
            try:
                col_qc_diag = _resolve_column_name(
                    df_diag, cfg_diag.col_qc, "col_qc"
                )
                col_qst_diag = _resolve_column_name(
                    df_diag, cfg_diag.col_qst, "col_qst"
                )
            except Exception:
                c.showPage()
                continue

            # Convertir les unites pour le graphique
            unit_qc_diag = "MPa"
            unit_qst_diag = "kN"
            if raw_data_manager:
                unit_qc_diag = raw_data_manager.get_unit(
                    file_path_diag, "unit_qc"
                )
                unit_qst_diag = raw_data_manager.get_unit(
                    file_path_diag, "unit_qst"
                )

            tip_area_diag = (
                settings_manager.get("unites", "tip_area_cm2") or 10.0
            )

            df_plot = df_diag.copy()
            qc_int = qc_to_internal(
                df_diag[col_qc_diag].values, unit_qc_diag, tip_area_diag,
            )
            qst_int = qst_to_internal(
                df_diag[col_qst_diag].values, unit_qst_diag,
            )
            qc_plt, qst_plt, _, _ = internal_to_plot(
                qc_int, qst_int, _plot_pair, tip_area_diag,
            )
            df_plot[col_qc_diag] = qc_plt
            df_plot[col_qst_diag] = qst_plt

            # ── En-tete (identique aux pages tableaux) ──
            location = essai.get("location", "").strip()
            street = essai.get("street", "").strip()
            test_id = essai.get("test", "").strip()
            dossier = essai.get("job", "").strip()
            date_str = _format_date_for_pdf(essai.get("date", ""))

            prof_atteinte = essai.get("prof_atteinte")
            prof_atteinte_str = (
                f"{prof_atteinte:.2f} m".replace(".", ",")
                if prof_atteinte is not None else ""
            )

            cote_depart_val = (cotes or {}).get(file_path_diag, 0.0)
            cote_depart_str = f"{cote_depart_val:.2f} m".replace(".", ",")

            top_y = PAGE_H - TOP_MARGIN
            y = top_y

            c.setFont(font_bold, FONT_SIZE_TITLE)
            c.drawString(LABELS_LEFT, y, location.upper() if location else "")
            y -= 14

            c.setFont(font_bold, FONT_SIZE_SUBTITLE)
            c.drawString(LABELS_LEFT, y, street)
            y -= 18

            c.setFont(font_normal, FONT_SIZE_SMALL)
            c.drawString(LABELS_LEFT, y, company_line1)
            y -= 8
            c.drawString(LABELS_LEFT, y, company_line2)
            y -= 8

            # Logo
            text_visual_top = top_y + FONT_SIZE_TITLE * 0.75
            text_visual_bottom = (
                top_y - 17 - 18 - FONT_SIZE_TEST_TYPE * 0.25
            )
            text_center_y = (text_visual_top + text_visual_bottom) / 2

            if logo_path:
                logo_y = (
                    text_center_y - LOGO_HEIGHT_PX / 2 + LOGO_V_OFFSET
                )
                try:
                    c.drawImage(
                        logo_path, LOGO_LEFT, logo_y,
                        width=LOGO_WIDTH_PX, height=LOGO_HEIGHT_PX,
                        preserveAspectRatio=True, anchor='sw',
                        mask='auto',
                    )
                except Exception as e:
                    logger.warning(
                        "Impossible de charger le logo PDF: %s", e,
                    )

            # Bloc metadonnees (droite)
            meta_y = top_y + 2
            for label, value, dy in [
                ("Dossier:", dossier, 0),
                ("Date:", date_str, -13),
                ("Prof. Atteinte:", prof_atteinte_str, -18),
                ("Cote de départ:", cote_depart_str, -13),
            ]:
                meta_y += dy
                c.setFont(font_normal, FONT_SIZE_META)
                c.drawString(META_X_LABEL, meta_y, label)
                c.setFont(font_bold, FONT_SIZE_META)
                c.drawRightString(META_X_VALUE, meta_y, value)

            header_end_y = y

            # ── Pied de page diagramme (observations + machine + légende) ──
            diag_footer_h = _draw_diagram_footer(
                c, LEFT_MARGIN, BOTTOM_MARGIN, TABLE_WIDTH,
                font_normal, font_bold, observations, file_path_diag,
                essai, settings_manager, _plot_pair,
            )

            # ── Dimensions du diagramme ──
            diagram_top_y = header_end_y - 20
            diagram_bottom_y = (
                BOTTOM_MARGIN + diag_footer_h + 1 * mm
            )
            diagram_width_pt = TABLE_WIDTH
            diagram_height_pt = diagram_top_y - diagram_bottom_y

            if diagram_height_pt > 0:
                diagram_w_in = diagram_width_pt / 72.0
                diagram_h_in = diagram_height_pt / 72.0

                plot_config = CPTPlotConfig(
                    show_titles=False,
                    plot_pair=_plot_pair,
                    resample_interval=0.20,
                    figure_dpi=300,
                    figure_width=diagram_w_in,
                    figure_height=diagram_h_in,
                    #adjust_left=0.04,    # Réduit la marge gauche au strict nécessaire pour les labels
                    #adjust_right=0.95,   # Étire le graphique presque jusqu'au bord droit
                    #adjust_top=0.92,     # Réduit l'espace vide en haut (puisqu'il n'y a pas de titre)
                    adjust_bottom=0.03,   # Réduit l'espace vide en bas
                    qc_color="black",
                    qst_color="black"
                )

                try:
                    fig_diag, _ax1, _ax2 = plot_cpt(df_plot, plot_config)
                except ValueError:
                    # Fallback sans reechantillonnage si le pas des
                    # donnees ne le permet pas
                    plot_config_nr = CPTPlotConfig(
                        show_titles=False,
                        plot_pair=_plot_pair,
                        figure_dpi=300,
                        figure_width=diagram_w_in,
                        figure_height=diagram_h_in,
                        #adjust_left=0.04,    # Réduit la marge gauche au strict nécessaire pour les labels
                        #adjust_right=0.95,   # Étire le graphique presque jusqu'au bord droit
                        #adjust_top=0.92,     # Réduit l'espace vide en haut (puisqu'il n'y a pas de titre)
                        adjust_bottom=0.03,   # Réduit l'espace vide en bas
                        qc_color="black",
                        qst_color="black"
                    )
                    try:
                        fig_diag, _ax1, _ax2 = plot_cpt(
                            df_plot, plot_config_nr,
                        )
                    except Exception as exc:
                        logger.error(
                            "Erreur diagramme (fallback) %s: %s",
                            file_path_diag, exc,
                        )
                        c.showPage()
                        continue

                # Extraction des coordonnées ax2 avant fermeture de la figure
                ax2_pos = _ax2.get_position()
                ax2_spine_lw = _ax2.spines['top'].get_linewidth()

                try:
                    img_data = _render_figure_to_image(fig_diag, dpi=300)
                    img_reader = ImageReader(img_data)
                    c.drawImage(
                        img_reader, LEFT_MARGIN, diagram_bottom_y,
                        width=diagram_width_pt,
                        height=diagram_height_pt,
                        preserveAspectRatio=True, anchor='sw',
                    )

                    # Boîte ouverte en bas (3 côtés) au-dessus du diagramme
                    box_x_left = LEFT_MARGIN + ax2_pos.x0 * diagram_width_pt
                    box_x_right = LEFT_MARGIN + ax2_pos.x1 * diagram_width_pt
                    box_y_bottom = (
                        diagram_bottom_y + ax2_pos.y1 * diagram_height_pt
                    )
                    box_y_top = box_y_bottom + 18 * mm

                    c.saveState()
                    c.setStrokeColor(black)
                    c.setLineWidth(ax2_spine_lw)
                    p = c.beginPath()
                    p.moveTo(box_x_left, box_y_bottom)
                    p.lineTo(box_x_left, box_y_top)
                    p.lineTo(box_x_right, box_y_top)
                    p.lineTo(box_x_right, box_y_bottom)
                    c.drawPath(p, stroke=1, fill=0)
                    c.restoreState()

                    # Titre et identifiant d'essai dans la boîte
                    box_title = "SONDAGE AU PENETROMETRE STATIQUE (CPT)"
                    title_fs = FONT_SIZE_TEST_TYPE + 1
                    id_fs = 14
                    text_x = box_x_left + 5 * mm
                    text_y = box_y_top - 2 * mm - title_fs

                    c.setFont(font_bold, title_fs)
                    title_w = pm.stringWidth(
                        box_title, font_bold, title_fs,
                    )
                    c.drawString(text_x, text_y, box_title)
                    c.setFont(font_bold, id_fs)
                    c.drawString(text_x + title_w + 15, text_y, test_id)

                except Exception as exc:
                    logger.error(
                        "Erreur rendu diagramme %s: %s",
                        file_path_diag, exc,
                    )

            c.showPage()

        # ══════════════════════════════════════════════════════════════
        # Pages tableaux (une ou plusieurs pages par essai)
        # ══════════════════════════════════════════════════════════════
        for essai_idx, essai in enumerate(job_essais):
            current += 1
            test_name = essai.get("test", "").strip() or "Essai"

            if progress_callback:
                progress_callback(
                    current, total_essais,
                    f"PDF {job_number} - {test_name}"
                )

            # Calculer les donnees
            data = _compute_essai_data(
                essai, settings_manager, cleaning_entries,
                raw_data_manager, cotes, observations,
            )

            alpha_essai = essai.get("alpha", 1.5)
            alpha_str = (
                f"{alpha_essai:.1f}".replace(".", ",")
                if alpha_essai != int(alpha_essai)
                else str(int(alpha_essai))
            )

            # Metadonnees pour l'en-tete
            location = essai.get("location", "").strip()
            street = essai.get("street", "").strip()
            test_id = essai.get("test", "").strip()
            dossier = essai.get("job", "").strip()
            date_str = _format_date_for_pdf(essai.get("date", ""))

            prof_atteinte = essai.get("prof_atteinte")
            prof_atteinte_str = (
                f"{prof_atteinte:.2f} m".replace(".", ",")
                if prof_atteinte is not None else ""
            )

            cote_depart_val = (cotes or {}).get(essai.get("file_path", ""), 0.0)
            cote_depart_str = f"{cote_depart_val:.2f} m".replace(".", ",")

            # ── Dessiner l'en-tete ──
            top_y = PAGE_H - TOP_MARGIN
            y = top_y

            c.setFont(font_bold, FONT_SIZE_TITLE)
            c.drawString(LABELS_LEFT, y, location.upper() if location else "")
            y -= 14

            c.setFont(font_bold, FONT_SIZE_SUBTITLE)
            c.drawString(LABELS_LEFT, y, street)
            y -= 18

            c.setFont(font_bold, FONT_SIZE_TEST_TYPE)
            test_type_text = "SONDAGE AU PENETROMETRE STATIQUE"
            c.drawString(LABELS_LEFT, y, test_type_text)
            type_w = pm.stringWidth(test_type_text, font_bold, FONT_SIZE_TEST_TYPE)
            c.setFont(font_bold, 13)
            c.drawString(LABELS_LEFT + type_w + 15, y, test_id)
            y -= 13

            c.setFont(font_normal, FONT_SIZE_SMALL)
            c.drawString(LABELS_LEFT, y, company_line1)
            y -= 8
            c.drawString(LABELS_LEFT, y, company_line2)
            y -= 8

            # Logo
            text_visual_top = top_y + FONT_SIZE_TITLE * 0.75
            text_visual_bottom = top_y - 17 - 18 - FONT_SIZE_TEST_TYPE * 0.25
            text_center_y = (text_visual_top + text_visual_bottom) / 2

            if logo_path:
                logo_y = text_center_y - LOGO_HEIGHT_PX / 2 + LOGO_V_OFFSET
                try:
                    c.drawImage(logo_path, LOGO_LEFT, logo_y,
                                width=LOGO_WIDTH_PX, height=LOGO_HEIGHT_PX,
                                preserveAspectRatio=True, anchor='sw',
                                mask='auto')
                except Exception as e:
                    logger.warning("Impossible de charger le logo PDF: %s", e)

            # Bloc metadonnees (droite)
            meta_y = top_y + 2
            for label, value, dy in [
                ("Dossier:", dossier, 0),
                ("Date:", date_str, -13),
                ("Prof. Atteinte:", prof_atteinte_str, -18),
                ("Cote de départ:", cote_depart_str, -13),
            ]:
                meta_y += dy
                c.setFont(font_normal, FONT_SIZE_META)
                c.drawString(META_X_LABEL, meta_y, label)
                c.setFont(font_bold, FONT_SIZE_META)
                c.drawRightString(META_X_VALUE, meta_y, value)

            header_end_y = y

            # ── En-tete du tableau ──
            table_start_y = header_end_y - 8
            n_cols = len(COL_WIDTHS)
            total_h = HEADER_ROW_HEIGHT + UNIT_ROW_HEIGHT
            block_top = table_start_y
            block_bottom = table_start_y - total_h
            row_sep_y = table_start_y - HEADER_ROW_HEIGHT

            c.setStrokeColor(black)
            c.setLineWidth(0.8)
            c.rect(LEFT_MARGIN, block_bottom, TABLE_WIDTH, total_h,
                   stroke=1, fill=0)

            for i in range(1, n_cols):
                c.line(col_x(i), block_top, col_x(i), block_bottom)

            # Titres des colonnes PDF :
            # Prof, Cote, qc, q'0, phi', phi_u, Nq, Ngamma, Padm_B1, Padm_B2, C
            largeur_semelle_1 = (
                settings_manager.get("parametres_calcul", "largeur_semelle_fondation_1")
                or 0.6
            )
            largeur_semelle_2 = (
                settings_manager.get("parametres_calcul", "largeur_semelle_fondation_2")
                or 1.5
            )
            b1_cm = int(round(largeur_semelle_1 * 100))
            b2_cm = int(round(largeur_semelle_2 * 100))

            labels = [
                [("Prof", "bold")],
                [("Cote", "bold")],
                [("q", "bold"), ("c", "sub")],
                [("q'", "bold"), ("o", "sub")],
                [(_PHI + "'", "bold")],
                [(_PHI, "bold"), ("u", "sub")],
                [("N", "bold"), ("q", "sub")],
                [("N", "bold"), ("\u03B3", "sub")],
                [("P", "bold"), (f"adm,{b1_cm}", "sub")],
                [("P", "bold"), (f"adm,{b2_cm}", "sub")],
                [("C", "bold")],
            ]

            ARIAL_COL_INDICES = {4, 5, 6, 7}

            text_y = row_sep_y + (HEADER_ROW_HEIGHT - FONT_SIZE_COL_HEADER) / 2 + 1
            for i, frags in enumerate(labels):
                if i in ARIAL_COL_INDICES:
                    _pdf_draw_text_with_sub_super(
                        c, col_x(i), text_y, frags, font_arial,
                        FONT_SIZE_COL_HEADER, "center", COL_WIDTHS[i],
                        font_bold=font_arial_bold)
                else:
                    _pdf_draw_text_with_sub_super(
                        c, col_x(i), text_y, frags, font_bold,
                        FONT_SIZE_COL_HEADER, "center", COL_WIDTHS[i])

            # Unites
            units = [
                [("[m]", "normal")],
                [("[m]", "normal")],
                [("[kg/cm", "normal"), ("2", "super"), ("]", "normal")],
                [("[kg/cm", "normal"), ("2", "super"), ("]", "normal")],
                [("[°]", "normal")],
                [("[°]", "normal")],
                [("[/]", "normal")],
                [("[/]", "normal")],
                [("[kg/cm", "normal"), ("2", "super"), ("]", "normal")],
                [("[kg/cm", "normal"), ("2", "super"), ("]", "normal")],
                [(f"[/] \u03B1={alpha_str}", "normal")],
            ]

            text_y = block_bottom + (UNIT_ROW_HEIGHT - FONT_SIZE_COL_UNIT) / 2 + 1
            for i, frags in enumerate(units):
                _pdf_draw_text_with_sub_super(
                    c, col_x(i), text_y, frags, font_normal,
                    FONT_SIZE_COL_UNIT, "center", COL_WIDTHS[i])

            header_block_bottom = block_bottom

            # ── Donnees du tableau ──
            if data is not None:
                resampled = data["resampled"]
                n_data = min(len(resampled), _PDF_MAX_DATA_ROWS)
            else:
                n_data = 0

            if n_data > 0:
                block_height = n_data * ROW_HEIGHT
                data_top = header_block_bottom
                data_bottom = data_top - block_height

                c.setStrokeColor(black)
                c.setLineWidth(0.8)
                c.rect(LEFT_MARGIN, data_bottom, TABLE_WIDTH, block_height,
                       stroke=1, fill=0)

                for i in range(1, n_cols):
                    c.line(col_x(i), data_top, col_x(i), data_bottom)

                # Mapping colonnes PDF :
                # 0:Prof, 1:Cote, 2:qc, 3:q'0, 4:phi', 5:phi_u,
                # 6:Nq, 7:Ng, 8:Padm1, 9:Padm2, 10:C
                #
                # Formats PDF :
                # - Prof (0), Cote (1), q'0 (3) : format 0.00
                # - Toutes les autres : format 0.0
                for row_idx in range(n_data):
                    row_y = data_top - (row_idx + 1) * ROW_HEIGHT
                    text_y_row = row_y + (ROW_HEIGHT - FONT_SIZE_DATA) / 2 + 1

                    depth_val = data["resampled"][row_idx]
                    cote_val = data["cote_depart"] - depth_val
                    q0_val = data["q0_values"][row_idx]
                    qc_val = data["qc_values"][row_idx]
                    phi_p = data["phi_prime_values"][row_idx]
                    phi_u_val = data["phi_u_values"][row_idx]
                    nq_val = data["nq_values"][row_idx]
                    ng_val = data["ng_values"][row_idx]
                    padm1_val = data["padm1_values"][row_idx]
                    padm2_val = data["padm2_values"][row_idx]
                    coeff_c_val = data["coeff_c_values"][row_idx]

                    # Construire les valeurs de la ligne
                    # Colonnes PDF : Prof, Cote, qc, q'0, phi', phi_u, Nq, Ng, Padm1, Padm2, C
                    row_values = [
                        (depth_val, "0.00"),
                        (cote_val, "0.00"),
                        (qc_val, "0.0"),
                        (q0_val, "0.00"),
                        (phi_p, "0.0"),
                        (phi_u_val, "0.0"),
                        (nq_val, "0.0"),
                        (ng_val, "0.0"),
                        (padm1_val, "0.0"),
                        (padm2_val, "0.0"),
                        (coeff_c_val, "0.0"),
                    ]

                    for col_idx, (val, fmt) in enumerate(row_values):
                        if val is not None:
                            if fmt == "0.00":
                                text = f"{val:.2f}".replace(".", ",")
                            else:
                                text = f"{val:.1f}".replace(".", ",")
                            c.setFont(font_normal, FONT_SIZE_DATA)
                            c.drawRightString(
                                col_x(col_idx) + COL_WIDTHS[col_idx] - 3,
                                text_y_row, text)

            # ── Pied de page ──
            c.setStrokeColor(black)
            c.setLineWidth(0.8)
            c.rect(LEFT_MARGIN, BOTTOM_MARGIN, TABLE_WIDTH, FOOTER_BOX_HEIGHT,
                   stroke=1, fill=0)

            line_y = BOTTOM_MARGIN + FOOTER_BOX_HEIGHT - 13
            _pdf_draw_footer_line_with_m3(
                c, LEFT_MARGIN + 5, line_y,
                f"Masse volumique du sol saturé: {int(rho_sat)} ",
                font_normal, FONT_SIZE_FOOTER)

            line_y -= 12
            _pdf_draw_footer_line_with_m3(
                c, LEFT_MARGIN + 5, line_y,
                f"Masse volumique du sol: {int(rho_sec)} ",
                font_normal, FONT_SIZE_FOOTER)

            line_y -= 12
            frags = [
                ("P", "bold"),
                ("adm,B", "sub"),
                (f" = pression admissible sous une semelle de B cm de largeur "
                 f"(avec un coefficient de sécurité égal à {int(coeff_securite)})",
                 "normal"),
            ]
            _pdf_draw_text_with_sub_super(
                c, LEFT_MARGIN + 5, line_y, frags, font_normal,
                FONT_SIZE_FOOTER, "left", TABLE_WIDTH)

            # Nouvelle page (sauf pour le dernier essai)
            if essai_idx < len(job_essais) - 1:
                c.showPage()

        # Sauvegarder le PDF
        try:
            c.save()
            generated_files[job_number] = filepath
            logger.info("Fichier PDF genere : %s", filepath)
        except OSError as exc:
            logger.error(
                "Erreur d'ecriture du fichier PDF %s : %s", filepath, exc
            )
            raise

    return generated_files
