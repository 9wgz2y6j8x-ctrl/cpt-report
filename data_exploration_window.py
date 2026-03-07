"""
data_exploration_window.py

Fenetre d'exploration des donnees brutes d'un sondage CPT.
Affiche le DataFrame complet (brut ou filtre) sous forme de tableau,
avec coloration, copie presse-papiers et export XLSX.
"""

import customtkinter as ctk
from tkinter import ttk, filedialog
import tkinter as tk
import pandas as pd
import os
from typing import Optional

from cpt_cleaning_view import CPTFileEntry, COLORS, FONTS


# ---------------------------------------------------------------------------
# Mapping des roles metier des colonnes
# ---------------------------------------------------------------------------

_COLUMN_ROLES = {
    "col_depth": ("Profondeur", "Axe vertical du sondage (en metres)"),
    "col_qc": ("Resistance de pointe qc", "Resistance a la penetration du cone"),
    "col_qst": ("Frottement lateral Qst", "Frottement sur le manchon du cone"),
}


class DataExplorationWindow(ctk.CTkToplevel):
    """Fenetre modale d'exploration des donnees d'un sondage CPT."""

    def __init__(self, parent, entry: CPTFileEntry, raw_data_manager):
        super().__init__(parent)
        self._entry = entry
        self._rdm = raw_data_manager
        self._show_filtered = False

        self.title(f"Exploration — {entry.filename}")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()

        # Dimensionner et centrer
        w, h = 960, 640
        self.geometry(f"{w}x{h}")
        self.minsize(640, 400)
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"+{max(x, 0)}+{max(y, 0)}")

        self._build_ui()
        self._populate_table()

    # ──────────────────────── Construction UI ────────────────────────

    def _build_ui(self):
        """Construit l'interface de la fenetre."""
        # ─── En-tete ───
        header = ctk.CTkFrame(self, fg_color="#0115B8", corner_radius=0, height=48)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header,
            text=f"Donnees — {self._entry.filename}",
            font=("Verdana", 16, "bold"),
            text_color="white",
        ).pack(side="left", padx=16, pady=8)

        # ─── Barre d'outils ───
        toolbar = ctk.CTkFrame(self, fg_color="#F5F7FA", corner_radius=0, height=44)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        # Toggle brut / filtre
        self._toggle_var = tk.BooleanVar(value=False)
        has_filtered = self._entry.df_filtered is not None
        toggle_state = "normal" if has_filtered else "disabled"

        self._toggle_label = ctk.CTkLabel(
            toolbar,
            text="Donnees brutes",
            font=("Verdana", 12, "bold"),
            text_color="#1A1A1A",
        )
        self._toggle_label.pack(side="left", padx=(16, 8))

        self._toggle_switch = ctk.CTkSwitch(
            toolbar,
            text="Filtre",
            font=("Verdana", 11),
            variable=self._toggle_var,
            command=self._on_toggle,
            onvalue=True,
            offvalue=False,
            width=48,
            progress_color="#16A34A" if has_filtered else "#D0D0D0",
            button_color="white",
            fg_color="#CBD5E1",
        )
        self._toggle_switch.pack(side="left", padx=4)
        if not has_filtered:
            self._toggle_switch.configure(state="disabled")
            self._no_filtered = True
        else:
            self._no_filtered = False

        # Separateur
        ctk.CTkFrame(toolbar, fg_color="#CBD5E1", width=1).pack(
            side="left", fill="y", padx=12, pady=8
        )

        # Bouton Copier
        self._btn_copy = ctk.CTkButton(
            toolbar,
            text="Copier",
            font=("Verdana", 12),
            fg_color="#E8EDF2",
            hover_color="#D0D8E4",
            text_color="#1A1A1A",
            corner_radius=6,
            width=80,
            height=30,
            command=self._on_copy,
        )
        self._btn_copy.pack(side="left", padx=4)

        # Bouton Exporter XLSX
        self._btn_export = ctk.CTkButton(
            toolbar,
            text="Exporter XLSX",
            font=("Verdana", 12),
            fg_color="#E8EDF2",
            hover_color="#D0D8E4",
            text_color="#1A1A1A",
            corner_radius=6,
            width=120,
            height=30,
            command=self._on_export_xlsx,
        )
        self._btn_export.pack(side="left", padx=4)

        # Compteur lignes
        self._row_count_label = ctk.CTkLabel(
            toolbar,
            text="",
            font=("Verdana", 11),
            text_color="#6B7280",
        )
        self._row_count_label.pack(side="right", padx=16)

        # ─── Legende des colonnes utilisees ───
        self._legend_frame = ctk.CTkFrame(self, fg_color="#FAFAFA", corner_radius=0)
        self._legend_frame.pack(fill="x", padx=0, pady=0)
        self._build_column_legend()

        # ─── Tableau ───
        table_frame = ctk.CTkFrame(self, fg_color="white", corner_radius=0)
        table_frame.pack(fill="both", expand=True, padx=0, pady=0)
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self._configure_table_style()

        self._table = ttk.Treeview(
            table_frame,
            show="headings",
            selectmode="extended",
            style="DataExplorer.Treeview",
        )
        self._table.grid(row=0, column=0, sticky="nsew")

        # Scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self._table.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self._table.xview)
        hsb.grid(row=1, column=0, sticky="ew")
        self._table.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Tags de coloration
        self._table.tag_configure("oddrow", background="#F3F3F3", foreground="#2E2E2E")
        self._table.tag_configure("evenrow", background="white", foreground="#2E2E2E")
        self._table.tag_configure("modified", background="#FFF3C4", foreground="#6D4C00")

    def _configure_table_style(self):
        """Configure le style du treeview de donnees."""
        style = ttk.Style()
        style.configure(
            "DataExplorer.Treeview",
            background="white",
            foreground="#2E2E2E",
            fieldbackground="white",
            font=("Verdana", 11),
            rowheight=28,
            borderwidth=0,
            relief="flat",
        )
        style.configure(
            "DataExplorer.Treeview.Heading",
            font=("Verdana", 11, "bold"),
            background="#0115B8",
            foreground="white",
            borderwidth=0,
            relief="flat",
        )
        style.map(
            "DataExplorer.Treeview.Heading",
            background=[("active", "#1976D2")],
        )

    def _build_column_legend(self):
        """Affiche la legende des colonnes utilisees par le programme."""
        for w in self._legend_frame.winfo_children():
            w.destroy()

        entry = self._entry
        col_map = {}
        if entry.col_depth:
            col_map[entry.col_depth] = _COLUMN_ROLES["col_depth"]
        if entry.col_qc:
            col_map[entry.col_qc] = _COLUMN_ROLES["col_qc"]
        if entry.col_qst:
            col_map[entry.col_qst] = _COLUMN_ROLES["col_qst"]

        if not col_map:
            return

        row = ctk.CTkFrame(self._legend_frame, fg_color="transparent")
        row.pack(fill="x", padx=16, pady=(6, 6))

        ctk.CTkLabel(
            row,
            text="Colonnes utilisees :",
            font=("Verdana", 11, "bold"),
            text_color="#374151",
        ).pack(side="left", padx=(0, 10))

        for col_name, (role_name, role_desc) in col_map.items():
            chip = ctk.CTkFrame(row, fg_color="#E0E7FF", corner_radius=6)
            chip.pack(side="left", padx=3, pady=2)
            ctk.CTkLabel(
                chip,
                text=f"{col_name}",
                font=("Verdana", 10, "bold"),
                text_color="#3730A3",
            ).pack(side="left", padx=(8, 2), pady=2)
            ctk.CTkLabel(
                chip,
                text=f"= {role_name}",
                font=("Verdana", 10),
                text_color="#6366F1",
            ).pack(side="left", padx=(0, 8), pady=2)

    # ──────────────────────── Remplissage du tableau ────────────────────────

    def _get_current_df(self) -> Optional[pd.DataFrame]:
        """Retourne le DataFrame courant selon le toggle."""
        if self._show_filtered:
            return self._entry.df_filtered
        return self._entry.df_raw

    def _populate_table(self):
        """Remplit le treeview avec les donnees du DataFrame."""
        df = self._get_current_df()
        if df is None:
            self._row_count_label.configure(text="Aucune donnee")
            return

        # Nettoyer le tableau
        self._table.delete(*self._table.get_children())

        # Configurer les colonnes a partir du DataFrame reel
        cols = list(df.columns)
        self._table["columns"] = cols
        for col_name in cols:
            self._table.heading(col_name, text=col_name, anchor="center")
            self._table.column(col_name, width=110, minwidth=70, anchor="center")

        # Determiner les colonnes utilisees et les lignes modifiees
        used_cols = set()
        if self._entry.col_depth:
            used_cols.add(self._entry.col_depth)
        if self._entry.col_qc:
            used_cols.add(self._entry.col_qc)
        if self._entry.col_qst:
            used_cols.add(self._entry.col_qst)

        # Si on est en mode filtre, detecter les lignes modifiees
        modified_rows = set()
        if self._show_filtered and self._entry.df_raw is not None and self._entry.df_filtered is not None:
            df_raw = self._entry.df_raw
            df_filt = self._entry.df_filtered
            if len(df_raw) == len(df_filt):
                for idx in range(len(df_raw)):
                    for col_name in used_cols:
                        if col_name in df_raw.columns and col_name in df_filt.columns:
                            raw_val = df_raw.iloc[idx][col_name]
                            filt_val = df_filt.iloc[idx][col_name]
                            try:
                                if pd.notna(raw_val) and pd.notna(filt_val) and raw_val != filt_val:
                                    modified_rows.add(idx)
                                    break
                            except (TypeError, ValueError):
                                pass

        # Remplir les lignes
        for i, (idx, row) in enumerate(df.iterrows()):
            values = []
            for col_name in cols:
                val = row[col_name]
                if pd.isna(val):
                    values.append("")
                elif isinstance(val, float):
                    values.append(f"{val:.4f}" if abs(val) < 1000 else f"{val:.2f}")
                else:
                    values.append(str(val))

            if i in modified_rows:
                tag = "modified"
            elif i % 2 == 0:
                tag = "evenrow"
            else:
                tag = "oddrow"

            self._table.insert("", "end", values=tuple(values), tags=(tag,))

        self._row_count_label.configure(
            text=f"{len(df)} lignes x {len(cols)} colonnes"
        )

    # ──────────────────────── Actions ────────────────────────

    def _on_toggle(self):
        """Bascule entre donnees brutes et filtrees."""
        self._show_filtered = self._toggle_var.get()
        if self._show_filtered:
            self._toggle_label.configure(text="Donnees filtrees", text_color="#16A34A")
        else:
            self._toggle_label.configure(text="Donnees brutes", text_color="#1A1A1A")
        self._populate_table()

    def _on_copy(self):
        """Copie le tableau affiche dans le presse-papiers (format TSV)."""
        df = self._get_current_df()
        if df is None:
            return

        text = df.to_csv(sep="\t", index=False)
        self.clipboard_clear()
        self.clipboard_append(text)

        # Retour visuel
        original_text = self._btn_copy.cget("text")
        self._btn_copy.configure(text="Copie !", fg_color="#16A34A", text_color="white")
        self.after(1500, lambda: self._btn_copy.configure(
            text=original_text, fg_color="#E8EDF2", text_color="#1A1A1A"
        ))

    def _on_export_xlsx(self):
        """Exporte le tableau affiche en fichier XLSX."""
        df = self._get_current_df()
        if df is None:
            return

        suffix = "_filtre" if self._show_filtered else "_brut"
        default_name = f"{os.path.splitext(self._entry.filename)[0]}{suffix}.xlsx"

        filepath = filedialog.asksaveasfilename(
            parent=self,
            title="Exporter en XLSX",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Tous", "*.*")],
            initialfile=default_name,
        )
        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Donnees CPT"

            # En-tetes
            header_font = Font(name="Verdana", size=10, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0115B8", end_color="0115B8", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )

            cols = list(df.columns)
            for c, col_name in enumerate(cols, 1):
                cell = ws.cell(row=1, column=c, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Donnees
            data_font = Font(name="Verdana", size=10)
            modified_fill = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")

            # Detecter les lignes modifiees pour le XLSX aussi
            modified_rows = set()
            if self._show_filtered and self._entry.df_raw is not None:
                df_raw = self._entry.df_raw
                used_cols = set()
                if self._entry.col_depth:
                    used_cols.add(self._entry.col_depth)
                if self._entry.col_qc:
                    used_cols.add(self._entry.col_qc)
                if self._entry.col_qst:
                    used_cols.add(self._entry.col_qst)
                if len(df_raw) == len(df):
                    for idx in range(len(df)):
                        for cn in used_cols:
                            if cn in df_raw.columns and cn in df.columns:
                                try:
                                    rv = df_raw.iloc[idx][cn]
                                    fv = df.iloc[idx][cn]
                                    if pd.notna(rv) and pd.notna(fv) and rv != fv:
                                        modified_rows.add(idx)
                                        break
                                except (TypeError, ValueError):
                                    pass

            for r, (idx, row) in enumerate(df.iterrows(), 2):
                for c, col_name in enumerate(cols, 1):
                    val = row[col_name]
                    cell = ws.cell(row=r, column=c, value=val if pd.notna(val) else None)
                    cell.font = data_font
                    cell.border = thin_border
                    if (r - 2) in modified_rows:
                        cell.fill = modified_fill

            # Ajuster la largeur des colonnes
            from openpyxl.utils import get_column_letter
            for c, col_name in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(c)].width = max(
                    len(str(col_name)) + 4, 14
                )

            wb.save(filepath)

            # Retour visuel
            original_text = self._btn_export.cget("text")
            self._btn_export.configure(text="Exporte !", fg_color="#16A34A", text_color="white")
            self.after(1500, lambda: self._btn_export.configure(
                text=original_text, fg_color="#E8EDF2", text_color="#1A1A1A"
            ))

        except Exception as exc:
            self._btn_export.configure(text=f"Erreur", fg_color="#C62828", text_color="white")
            self.after(2000, lambda: self._btn_export.configure(
                text="Exporter XLSX", fg_color="#E8EDF2", text_color="#1A1A1A"
            ))
